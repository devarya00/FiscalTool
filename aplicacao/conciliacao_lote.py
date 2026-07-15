"""Conciliação em lote multi-CNPJ: Fiscal (mensal) x Contábil (anual), lidos
de Excel — ferramenta separada do app desktop (que faz 1 empresa por vez via
upload de PDF).

SQLite :memory: faz a CARGA e o CRUZAMENTO por CNPJ (JOIN): detecta CNPJ que
existe só de um lado (alerta de nível de portfólio — distinto do P6.ORFAOS,
que alerta acumulador individual dentro de um CNPJ já casado) e calcula os
totais rápidos que viram total_fiscal/total_contabil no Excel de saída. A
correção GRANULAR (agrupamento N-para-N, depreciação fuzzy, órfãos) roda no
motor Python já validado (dominio/regras/*), por CNPJ — mesmo motor do app
desktop, reaproveitado sem alteração.

Cada CNPJ é tratado como unidade independente: nenhum estado é compartilhado
entre eles, então o custo cresce linearmente com o tamanho do portfólio.
"""
from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from dominio.apontamento import Apontamento, Severidade
from dominio.modelos import Balancete, Contexto, RelatorioFiscal
from dominio.motor import MotorRegras
from dominio.regras.depreciacao import P8Depreciacao
from dominio.regras.grupos import P4GrupoAmarracao
from dominio.regras.orfaos import P6Orfaos
from infra.excel_parser import ler_contabil_excel, ler_fiscal_excel

_CONFIG_PADRAO = Path(__file__).resolve().parent.parent / "config" / "regras.yaml"
_TOLERANCIA_PADRAO = Decimal("0.01")

_COR_STATUS = {"OK": "C9E7C9", "DIVERGENTE": "F2B6B3"}
_COR_SEVERIDADE = {
    Severidade.IMPEDITIVO: "7A0000", Severidade.CRITICO: "F2B6B3",
    Severidade.ALERTA: "FBE1B0", Severidade.OK: "C9E7C9",
}


@dataclass
class LinhaResumo:
    cnpj: str
    total_fiscal: Decimal
    total_contabil: Decimal
    status: str
    diferenca: Decimal
    apontamentos: list[Apontamento]


def _motor_lote() -> MotorRegras:
    """Subconjunto de regras relevante para a conciliação em lote pedida:
    agrupamento N-para-N (cobre Frete/Impostos/Imobilizado via config),
    depreciação com fuzzy matching e órfãos. Não inclui P1/P2/Folha/Simples/
    Receita×Custo — são do fluxo detalhado do app desktop, fora deste escopo."""
    return MotorRegras([P4GrupoAmarracao(), P8Depreciacao(), P6Orfaos()])


def _carregar_config(config_path: str | Path) -> dict[str, Any]:
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _preparar_sqlite(
    fiscais: dict[str, RelatorioFiscal], contabeis: dict[str, Balancete],
) -> sqlite3.Connection:
    """Carrega os dados já normalizados (Decimal, via to_decimal nos parsers)
    em duas tabelas SQLite em memória. A precisão financeira fica com o motor
    Python (Decimal) — o SQLite aqui só cruza/soma para o resumo de
    portfólio, então converter para REAL não afeta a auditoria em si."""
    conn = sqlite3.connect(":memory:")
    conn.executescript("""
        CREATE TABLE fiscal (cnpj TEXT, codigo TEXT, valor REAL);
        CREATE TABLE contabil (cnpj TEXT, codigo INTEGER, debito REAL, credito REAL);
        CREATE INDEX idx_fiscal_cnpj ON fiscal(cnpj);
        CREATE INDEX idx_contabil_cnpj ON contabil(cnpj);
    """)
    conn.executemany(
        "INSERT INTO fiscal (cnpj, codigo, valor) VALUES (?, ?, ?)",
        [
            (cnpj, ac.codigo, float(ac.valor_contabil))
            for cnpj, rel in fiscais.items() for ac in rel.todos()
        ],
    )
    conn.executemany(
        "INSERT INTO contabil (cnpj, codigo, debito, credito) VALUES (?, ?, ?, ?)",
        [
            (cnpj, conta.codigo, float(conta.debito), float(conta.credito))
            for cnpj, bal in contabeis.items() for conta in bal.contas
        ],
    )
    conn.commit()
    return conn


def _universo_cnpjs(conn: sqlite3.Connection) -> tuple[set[str], set[str], set[str]]:
    """JOIN por cnpj: quem existe nos dois lados (segue para o motor de
    regras), só no fiscal, só no contábil (ambos viram alerta de nível de
    portfólio — "CNPJ não encontrado no outro lado")."""
    ambos = {
        r[0] for r in conn.execute(
            "SELECT DISTINCT f.cnpj FROM fiscal f JOIN contabil c ON f.cnpj = c.cnpj"
        )
    }
    apenas_fiscal = {
        r[0] for r in conn.execute(
            "SELECT DISTINCT f.cnpj FROM fiscal f LEFT JOIN contabil c ON f.cnpj = c.cnpj "
            "WHERE c.cnpj IS NULL"
        )
    }
    apenas_contabil = {
        r[0] for r in conn.execute(
            "SELECT DISTINCT c.cnpj FROM contabil c LEFT JOIN fiscal f ON f.cnpj = c.cnpj "
            "WHERE f.cnpj IS NULL"
        )
    }
    return ambos, apenas_fiscal, apenas_contabil


def _totais_sql(conn: sqlite3.Connection, cnpj: str) -> tuple[Decimal, Decimal]:
    (total_fiscal,) = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) FROM fiscal WHERE cnpj = ?", (cnpj,),
    ).fetchone()
    (total_contabil,) = conn.execute(
        "SELECT COALESCE(SUM(debito), 0) + COALESCE(SUM(credito), 0) FROM contabil WHERE cnpj = ?", (cnpj,),
    ).fetchone()
    return Decimal(str(total_fiscal)), Decimal(str(total_contabil))


def conciliar_lote(
    caminho_fiscal: str, caminho_contabil: str, caminho_saida: str,
    config_path: str | Path = _CONFIG_PADRAO, tolerancia: Decimal = _TOLERANCIA_PADRAO,
) -> list[LinhaResumo]:
    config = _carregar_config(config_path)
    fiscais = ler_fiscal_excel(caminho_fiscal)
    contabeis = ler_contabil_excel(caminho_contabil)

    conn = _preparar_sqlite(fiscais, contabeis)
    try:
        ambos, apenas_fiscal, apenas_contabil = _universo_cnpjs(conn)
        motor = _motor_lote()
        resumos: list[LinhaResumo] = []

        for cnpj in sorted(ambos):
            total_fiscal, total_contabil = _totais_sql(conn, cnpj)
            ctx = Contexto(fiscal=fiscais[cnpj], balancete=contabeis[cnpj], config=config)
            apontamentos = motor.executar(ctx)

            diferenca = total_fiscal - total_contabil
            tem_problema = abs(diferenca) >= tolerancia or any(
                a.severidade in (Severidade.IMPEDITIVO, Severidade.CRITICO, Severidade.ALERTA)
                for a in apontamentos
            )
            resumos.append(LinhaResumo(
                cnpj=cnpj, total_fiscal=total_fiscal, total_contabil=total_contabil,
                status="DIVERGENTE" if tem_problema else "OK",
                diferenca=diferenca, apontamentos=apontamentos,
            ))

        for cnpj in sorted(apenas_fiscal):
            total_fiscal, _ = _totais_sql(conn, cnpj)
            resumos.append(LinhaResumo(
                cnpj=cnpj, total_fiscal=total_fiscal, total_contabil=Decimal("0"),
                status="DIVERGENTE", diferenca=total_fiscal,
                apontamentos=[Apontamento(
                    regra="G.CNPJ_AUSENTE", severidade=Severidade.ALERTA,
                    descricao=f"CNPJ {cnpj} presente no Fiscal mas não encontrado no Contábil",
                )],
            ))

        for cnpj in sorted(apenas_contabil):
            _, total_contabil = _totais_sql(conn, cnpj)
            resumos.append(LinhaResumo(
                cnpj=cnpj, total_fiscal=Decimal("0"), total_contabil=total_contabil,
                status="DIVERGENTE", diferenca=-total_contabil,
                apontamentos=[Apontamento(
                    regra="G.CNPJ_AUSENTE", severidade=Severidade.ALERTA,
                    descricao=f"CNPJ {cnpj} presente no Contábil mas não encontrado no Fiscal",
                )],
            ))
    finally:
        conn.close()

    resumos.sort(key=lambda r: r.cnpj)
    exportar_relatorio(resumos, caminho_saida)
    return resumos


def exportar_relatorio(resumos: list[LinhaResumo], caminho_saida: str) -> None:
    wb = Workbook()

    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo.append(["cnpj", "total_fiscal", "total_contabil", "status", "diferenca"])
    for cel in ws_resumo[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="333333")
    for r in resumos:
        ws_resumo.append([
            r.cnpj, float(r.total_fiscal), float(r.total_contabil), r.status, float(r.diferenca),
        ])
        cel_status = ws_resumo.cell(row=ws_resumo.max_row, column=4)
        cel_status.fill = PatternFill("solid", fgColor=_COR_STATUS[r.status])
        for col in (2, 3, 5):
            ws_resumo.cell(row=ws_resumo.max_row, column=col).number_format = "#,##0.00"
    for coluna in ws_resumo.columns:
        largura = max(len(str(c.value)) if c.value is not None else 0 for c in coluna) + 2
        ws_resumo.column_dimensions[coluna[0].column_letter].width = min(largura, 60)

    ws_detalhes = wb.create_sheet("Detalhes")
    ws_detalhes.append(["cnpj", "regra", "severidade", "descricao", "valor_fiscal", "valor_contabil", "diferenca"])
    for cel in ws_detalhes[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="333333")
    for r in resumos:
        for a in r.apontamentos:
            ws_detalhes.append([
                r.cnpj, a.regra, a.severidade.value.upper(), a.descricao,
                float(a.valor_fiscal) if a.valor_fiscal is not None else None,
                float(a.valor_contabil) if a.valor_contabil is not None else None,
                float(a.diferenca) if a.diferenca is not None else None,
            ])
            cor = _COR_SEVERIDADE.get(a.severidade)
            if cor:
                ws_detalhes.cell(row=ws_detalhes.max_row, column=3).fill = PatternFill("solid", fgColor=cor)
    for coluna in ws_detalhes.columns:
        largura = max(len(str(c.value)) if c.value is not None else 0 for c in coluna) + 2
        ws_detalhes.column_dimensions[coluna[0].column_letter].width = min(largura, 60)

    wb.save(caminho_saida)
