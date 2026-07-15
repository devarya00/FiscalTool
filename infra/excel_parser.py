"""Leitura de dados fiscais/contábeis tabulados em Excel, para conciliação em
lote multi-CNPJ (ver aplicacao/conciliacao_lote.py).

Reconstrói os MESMOS modelos de domínio que os parsers de PDF produzem
(RelatorioFiscal, Balancete) — o motor de regras (dominio/regras/*) roda
idêntico independente da origem do dado, sem saber que veio de uma planilha.

Fiscal é mensal: várias linhas por CNPJ (um mês cada) caem na mesma lista de
acumuladores — as regras de agrupamento (P4.GRUPO) já somam por código/padrão
sobre a lista inteira, então múltiplos meses não precisam de agregação
especial aqui. Contábil é anual: uma linha por conta por CNPJ.
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal

import openpyxl

from dominio.modelos import (
    Acumulador, Balancete, ContaBalancete, Grupo, Natureza, Periodo, RelatorioFiscal, Saldo, Secao,
)
from infra.normalizador import normalizar_cnpj, to_decimal

_MAPA_SECAO = {"entradas": Secao.ENTRADAS, "saidas": Secao.SAIDAS, "servicos": Secao.SERVICOS}
_MAPA_GRUPO = {"ativo": Grupo.ATIVO, "passivo": Grupo.PASSIVO, "pl": Grupo.PL, "resultado": Grupo.RESULTADO}
_MAPA_NATUREZA = {"d": Natureza.DEVEDOR, "c": Natureza.CREDOR}


class ColunaObrigatoriaAusente(Exception):
    """Planilha não tem uma coluna exigida — erro de configuração do arquivo
    de entrada, não do motor de regras."""


def _valor(bruto: object) -> Decimal:
    """Aceita tanto texto no formato contábil BR ('1.250,50') quanto números
    já convertidos pelo openpyxl (int/float) — a mesma limpeza de to_decimal
    (infra/normalizador.py) cobre os dois casos antes de qualquer cálculo."""
    if bruto is None:
        return Decimal("0")
    if isinstance(bruto, (int, float, Decimal)):
        return Decimal(str(bruto))
    return to_decimal(str(bruto))


def _data(bruto: object) -> date:
    if isinstance(bruto, date):
        return bruto
    dia, mes, ano = str(bruto).strip().split("/")
    return date(int(ano), int(mes), int(dia))


def _indice_colunas(ws, obrigatorias: list[str]) -> dict[str, int]:
    cabecalho = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {nome: i for i, nome in enumerate(cabecalho)}
    faltantes = [c for c in obrigatorias if c not in idx]
    if faltantes:
        raise ColunaObrigatoriaAusente(f"Coluna(s) obrigatória(s) ausente(s) na planilha: {faltantes}")
    return idx


def ler_fiscal_excel(caminho: str) -> dict[str, RelatorioFiscal]:
    """Colunas obrigatórias: cnpj, codigo, descricao, valor, secao
    (entradas|saidas|servicos), periodo_inicio, periodo_fim."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    idx = _indice_colunas(ws, ["cnpj", "codigo", "descricao", "valor", "secao", "periodo_inicio", "periodo_fim"])

    acumuladores_por_cnpj: dict[str, list[Acumulador]] = defaultdict(list)
    periodos_por_cnpj: dict[str, list[Periodo]] = defaultdict(list)

    for linha_idx, linha in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if linha[idx["cnpj"]] is None:
            continue
        cnpj = normalizar_cnpj(str(linha[idx["cnpj"]]))
        secao = _MAPA_SECAO.get(str(linha[idx["secao"]] or "").strip().lower(), Secao.ENTRADAS)
        ac = Acumulador(
            codigo=str(linha[idx["codigo"]]).strip(),
            descricao=str(linha[idx["descricao"]] or "").strip(),
            valor_contabil=_valor(linha[idx["valor"]]),
            secao=secao, pagina=1, linha=linha_idx,
        )
        acumuladores_por_cnpj[cnpj].append(ac)
        periodos_por_cnpj[cnpj].append(Periodo(
            _data(linha[idx["periodo_inicio"]]), _data(linha[idx["periodo_fim"]]),
        ))
    wb.close()

    resultado: dict[str, RelatorioFiscal] = {}
    for cnpj, acumuladores in acumuladores_por_cnpj.items():
        periodos = periodos_por_cnpj[cnpj]
        resultado[cnpj] = RelatorioFiscal(
            cnpj=cnpj,
            periodo=Periodo(min(p.inicio for p in periodos), max(p.fim for p in periodos)),
            entradas=[a for a in acumuladores if a.secao is Secao.ENTRADAS],
            saidas=[a for a in acumuladores if a.secao is Secao.SAIDAS],
            servicos=[a for a in acumuladores if a.secao is Secao.SERVICOS],
        )
    return resultado


def ler_contabil_excel(caminho: str) -> dict[str, Balancete]:
    """Colunas obrigatórias: cnpj, codigo, descricao, debito, credito,
    saldo_atual, periodo_inicio, periodo_fim. Opcionais (default se ausentes):
    grupo (default "ativo" — irrelevante para as regras deste motor de lote,
    que não usa P2.NATUREZA), natureza_atual (D/C), saldo_anterior,
    natureza_anterior."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    idx = _indice_colunas(
        ws, ["cnpj", "codigo", "descricao", "debito", "credito", "saldo_atual", "periodo_inicio", "periodo_fim"],
    )
    opcionais = ["grupo", "natureza_atual", "saldo_anterior", "natureza_anterior"]
    for nome in opcionais:
        if nome not in idx:
            idx[nome] = None

    contas_por_cnpj: dict[str, list[ContaBalancete]] = defaultdict(list)
    periodos_por_cnpj: dict[str, list[Periodo]] = defaultdict(list)

    for linha_idx, linha in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if linha[idx["cnpj"]] is None:
            continue
        cnpj = normalizar_cnpj(str(linha[idx["cnpj"]]))

        grupo_txt = str(linha[idx["grupo"]] or "ativo").strip().lower() if idx["grupo"] is not None else "ativo"
        natureza_txt = str(linha[idx["natureza_atual"]] or "").strip().lower() if idx["natureza_atual"] is not None else ""
        natureza_ant_txt = str(linha[idx["natureza_anterior"]] or "").strip().lower() if idx["natureza_anterior"] is not None else ""

        conta = ContaBalancete(
            codigo=int(linha[idx["codigo"]]),
            descricao=str(linha[idx["descricao"]] or "").strip(),
            grupo=_MAPA_GRUPO.get(grupo_txt, Grupo.ATIVO),
            saldo_anterior=Saldo(
                _valor(linha[idx["saldo_anterior"]]) if idx["saldo_anterior"] is not None else Decimal("0"),
                _MAPA_NATUREZA.get(natureza_ant_txt),
            ),
            debito=_valor(linha[idx["debito"]]),
            credito=_valor(linha[idx["credito"]]),
            saldo_atual=Saldo(_valor(linha[idx["saldo_atual"]]), _MAPA_NATUREZA.get(natureza_txt)),
            pagina=1, linha=linha_idx,
        )
        contas_por_cnpj[cnpj].append(conta)
        periodos_por_cnpj[cnpj].append(Periodo(
            _data(linha[idx["periodo_inicio"]]), _data(linha[idx["periodo_fim"]]),
        ))
    wb.close()

    resultado: dict[str, Balancete] = {}
    for cnpj, contas in contas_por_cnpj.items():
        periodos = periodos_por_cnpj[cnpj]
        resultado[cnpj] = Balancete(
            cnpj=cnpj,
            periodo=Periodo(min(p.inicio for p in periodos), max(p.fim for p in periodos)),
            contas=contas,
        )
    return resultado
