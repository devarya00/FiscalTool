"""Ponta a ponta da conciliação em lote: Excel -> SQLite (carga+JOIN por CNPJ)
-> motor Python (grupos_amarracao/depreciação/órfãos) -> Excel de saída."""
from __future__ import annotations

from decimal import Decimal

import openpyxl
from openpyxl import Workbook

from aplicacao.conciliacao_lote import conciliar_lote

_CABECALHO_FISCAL = ["cnpj", "codigo", "descricao", "valor", "secao", "periodo_inicio", "periodo_fim"]
_CABECALHO_CONTABIL = [
    "cnpj", "codigo", "descricao", "grupo", "debito", "credito", "saldo_atual",
    "natureza_atual", "periodo_inicio", "periodo_fim",
]

_CNPJ_OK = "11.111.111/0001-11"        # Frete bate exatamente -> OK
_CNPJ_DIVERGENTE = "22.222.222/0001-22"  # Frete não bate -> DIVERGENTE
_CNPJ_SO_FISCAL = "33.333.333/0001-33"   # existe só no fiscal
_CNPJ_SO_CONTABIL = "44.444.444/0001-44"  # existe só no contábil


def _gravar(caminho, cabecalho, linhas):
    wb = Workbook()
    ws = wb.active
    ws.append(cabecalho)
    for linha in linhas:
        ws.append(linha)
    wb.save(caminho)


def _preparar_arquivos(tmp_path):
    fiscal_path = tmp_path / "fiscal.xlsx"
    contabil_path = tmp_path / "contabil.xlsx"

    _gravar(fiscal_path, _CABECALHO_FISCAL, [
        [_CNPJ_OK, "300", "FRETE", "1.000,00", "entradas", "01/01/2025", "31/01/2025"],
        [_CNPJ_DIVERGENTE, "300", "FRETE", "1.000,00", "entradas", "01/01/2025", "31/01/2025"],
        [_CNPJ_SO_FISCAL, "300", "FRETE", "500,00", "entradas", "01/01/2025", "31/01/2025"],
    ])
    _gravar(contabil_path, _CABECALHO_CONTABIL, [
        [_CNPJ_OK, 688, "DESPESA COM FRETE", "resultado", "1.000,00", "0,00", "1.000,00",
         "D", "01/01/2025", "31/12/2025"],
        [_CNPJ_DIVERGENTE, 688, "DESPESA COM FRETE", "resultado", "700,00", "0,00", "700,00",
         "D", "01/01/2025", "31/12/2025"],
        [_CNPJ_SO_CONTABIL, 688, "DESPESA COM FRETE", "resultado", "200,00", "0,00", "200,00",
         "D", "01/01/2025", "31/12/2025"],
    ])
    return str(fiscal_path), str(contabil_path)


def test_cnpj_com_grupo_batendo_e_ok(tmp_path):
    fiscal_path, contabil_path = _preparar_arquivos(tmp_path)
    resumos = conciliar_lote(fiscal_path, contabil_path, str(tmp_path / "saida.xlsx"))

    por_cnpj = {r.cnpj: r for r in resumos}
    normalizado_ok = "11111111000111"
    assert por_cnpj[normalizado_ok].status == "OK"
    assert por_cnpj[normalizado_ok].diferenca == Decimal("0")


def test_cnpj_com_grupo_divergente(tmp_path):
    fiscal_path, contabil_path = _preparar_arquivos(tmp_path)
    resumos = conciliar_lote(fiscal_path, contabil_path, str(tmp_path / "saida.xlsx"))

    por_cnpj = {r.cnpj: r for r in resumos}
    normalizado_div = "22222222000122"
    linha = por_cnpj[normalizado_div]
    assert linha.status == "DIVERGENTE"
    assert linha.diferenca == Decimal("300.00")
    assert any(a.regra == "G.FRETE" for a in linha.apontamentos)


def test_cnpj_so_no_fiscal_vira_divergente_com_alerta_de_ausencia(tmp_path):
    fiscal_path, contabil_path = _preparar_arquivos(tmp_path)
    resumos = conciliar_lote(fiscal_path, contabil_path, str(tmp_path / "saida.xlsx"))

    por_cnpj = {r.cnpj: r for r in resumos}
    linha = por_cnpj["33333333000133"]
    assert linha.status == "DIVERGENTE"
    assert linha.total_contabil == Decimal("0")
    assert any(a.regra == "G.CNPJ_AUSENTE" for a in linha.apontamentos)


def test_cnpj_so_no_contabil_vira_divergente_com_alerta_de_ausencia(tmp_path):
    fiscal_path, contabil_path = _preparar_arquivos(tmp_path)
    resumos = conciliar_lote(fiscal_path, contabil_path, str(tmp_path / "saida.xlsx"))

    por_cnpj = {r.cnpj: r for r in resumos}
    linha = por_cnpj["44444444000144"]
    assert linha.status == "DIVERGENTE"
    assert linha.total_fiscal == Decimal("0")
    assert any(a.regra == "G.CNPJ_AUSENTE" for a in linha.apontamentos)


def test_cada_cnpj_e_tratado_como_unidade_independente(tmp_path):
    """CNPJ divergente não deve contaminar o resultado de outro CNPJ (nenhum
    estado global vaza entre eles)."""
    fiscal_path, contabil_path = _preparar_arquivos(tmp_path)
    resumos = conciliar_lote(fiscal_path, contabil_path, str(tmp_path / "saida.xlsx"))
    assert len(resumos) == 4  # 4 CNPJs distintos, cada um com seu próprio resultado


def test_tolerancia_evita_falso_positivo_por_centavo(tmp_path):
    """Config com tolerancia_centavos:1 — a diferença de meio centavo não deve
    virar CRITICO na regra de grupo nem DIVERGENTE no resumo do lote. O
    default do app desktop (tolerancia_centavos:0, igualdade exata) é uma
    decisão à parte e não é alterado aqui — o lote usa a config que receber."""
    fiscal_path = tmp_path / "fiscal.xlsx"
    contabil_path = tmp_path / "contabil.xlsx"
    config_path = tmp_path / "regras.yaml"
    config_path.write_text(
        "tolerancia_centavos: 1\n"
        "grupos_amarracao:\n"
        "  - id: G.FRETE\n"
        "    codigos_fiscais: [300]\n"
        "    contas_contabeis: [688]\n"
        "    coluna: DEBITO\n"
        "    modo: soma\n",
        encoding="utf-8",
    )
    _gravar(fiscal_path, _CABECALHO_FISCAL, [
        [_CNPJ_OK, "300", "FRETE", "1.000,00", "entradas", "01/01/2025", "31/01/2025"],
    ])
    _gravar(contabil_path, _CABECALHO_CONTABIL, [
        [_CNPJ_OK, 688, "DESPESA COM FRETE", "resultado", "1.000,005", "0,00", "1.000,00",
         "D", "01/01/2025", "31/12/2025"],
    ])

    resumos = conciliar_lote(
        str(fiscal_path), str(contabil_path), str(tmp_path / "saida.xlsx"), config_path=config_path,
    )
    assert resumos[0].status == "OK"


def test_exporta_excel_com_colunas_pedidas(tmp_path):
    fiscal_path, contabil_path = _preparar_arquivos(tmp_path)
    saida = tmp_path / "relatorio_conferencia.xlsx"
    conciliar_lote(fiscal_path, contabil_path, str(saida))

    assert saida.exists()
    wb = openpyxl.load_workbook(str(saida))
    assert "Resumo" in wb.sheetnames
    cabecalho = [c.value for c in next(wb["Resumo"].iter_rows(min_row=1, max_row=1))]
    assert cabecalho == ["cnpj", "total_fiscal", "total_contabil", "status", "diferenca"]
    assert wb["Resumo"].max_row == 5  # cabeçalho + 4 CNPJs
