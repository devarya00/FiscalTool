from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from aplicacao.conferencia_service import ResultadoConferencia
from dominio.apontamento import Apontamento, Severidade

_COR_SEVERIDADE = {
    Severidade.IMPEDITIVO: "7A0000",
    Severidade.CRITICO: "D9534F",
    Severidade.ALERTA: "F0AD4E",
    Severidade.OK: "5CB85C",
}

_CABECALHO = [
    "Regra", "Severidade", "Descrição", "Valor Fiscal", "Valor Contábil", "Diferença",
    "Coluna Esperada", "Coluna Encontrada", "Origem Fiscal (pág/linha)", "Origem Balancete (pág/linha)",
]


def _escrever_bloco(ws: Worksheet, apontamentos: list[Apontamento]) -> None:
    ws.append(_CABECALHO)
    for cel in ws[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="333333")

    for ap in apontamentos:
        origem_fiscal = f"{ap.origem_fiscal.pagina}/{ap.origem_fiscal.linha}" if ap.origem_fiscal else ""
        origem_balancete = (
            f"{ap.origem_balancete.pagina}/{ap.origem_balancete.linha}" if ap.origem_balancete else ""
        )
        ws.append([
            ap.regra, ap.severidade.value.upper(), ap.descricao,
            float(ap.valor_fiscal) if ap.valor_fiscal is not None else None,
            float(ap.valor_contabil) if ap.valor_contabil is not None else None,
            float(ap.diferenca) if ap.diferenca is not None else None,
            ap.coluna_esperada, ap.coluna_encontrada,
            origem_fiscal, origem_balancete,
        ])
        cor = _COR_SEVERIDADE.get(ap.severidade)
        if cor:
            ws.cell(row=ws.max_row, column=2).fill = PatternFill("solid", fgColor=cor)

    for coluna in ws.columns:
        largura = max(len(str(c.value)) if c.value is not None else 0 for c in coluna) + 2
        ws.column_dimensions[coluna[0].column_letter].width = min(largura, 60)


def exportar_xlsx(resultado: ResultadoConferencia, caminho: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    blocos = [
        ("Cabecalho", resultado.cabecalho()),
        ("Estrutural", resultado.estrutural()),
        ("Amarracao", resultado.amarracao()),
        ("Fluxo de Caixa", resultado.fluxo_de_caixa()),
        ("Nao Mapeados", resultado.orfaos()),
    ]
    for titulo, apontamentos in blocos:
        ws = wb.create_sheet(title=titulo[:31])
        _escrever_bloco(ws, apontamentos)

    wb.save(caminho)
